import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from io import BytesIO
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import plotly.express as px
import plotly.graph_objects as go
from plotly.offline import plot
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def register_fonts():
    """Регистрация шрифтов с поддержкой кириллицы"""
    try:
        # Получаем абсолютный путь к папке со скриптом
        script_dir = os.path.dirname(os.path.abspath(__file__))
        local_fonts_dir = os.path.join(script_dir, "fonts")
        
        # Список путей для поиска шрифтов (приоритет локальным)
        font_paths = [
            # Локальные шрифты (высший приоритет)
            local_fonts_dir,
            "./fonts/",
            # Системные пути Windows
            "C:/Windows/Fonts/",
            "C:/WINDOWS/Fonts/",
            # Пути для различных ОС
            "/usr/share/fonts/truetype/dejavu/",
            "/usr/share/fonts/TTF/",
            "/System/Library/Fonts/",
            "/Library/Fonts/",
        ]
        
        # Приоритетные шрифты с поддержкой кириллицы
        priority_fonts = [
            # DejaVu (лучшая поддержка кириллицы)
            ("DejaVuSans.ttf", "DejaVuSans"),
            ("DejaVuSans-Bold.ttf", "DejaVuSans-Bold"),
            # Liberation (хорошая альтернатива)
            ("LiberationSans-Regular.ttf", "LiberationSans"),
            ("LiberationSans-Bold.ttf", "LiberationSans-Bold"),
            # Системные шрифты Windows
            ("arial.ttf", "Arial"),
            ("arialbd.ttf", "Arial-Bold"),
            ("times.ttf", "Times"),
            ("timesbd.ttf", "Times-Bold"),
        ]
        
        registered_fonts = []
        
        # Попытка регистрации приоритетных шрифтов
        for font_file, font_name in priority_fonts:
            for base_path in font_paths:
                font_path = os.path.join(base_path, font_file)
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        registered_fonts.append(font_name)
                        print(f"✓ Зарегистрирован шрифт: {font_name} ({font_path})")
                        break
                    except Exception as e:
                        print(f"✗ Ошибка регистрации {font_name}: {e}")
                        continue
        
        # Fallback: регистрация Unicode CID шрифтов
        if not registered_fonts:
            try:
                # Попытка использовать встроенные Unicode шрифты ReportLab
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
                registered_fonts.append('HeiseiKakuGo-W5')
                print("✓ Зарегистрирован fallback Unicode шрифт: HeiseiKakuGo-W5")
            except:
                pass
        
        if registered_fonts:
            print(f"✓ Всего зарегистрировано шрифтов: {len(registered_fonts)}")
            return registered_fonts
        else:
            print("⚠ Не удалось зарегистрировать ни одного шрифта с поддержкой кириллицы")
            return []
            
    except Exception as e:
        print(f"✗ Ошибка при регистрации шрифтов: {e}")
        return []

class OrderAnalyzer:
    """Класс для анализа данных о заказах"""
    
    def __init__(self, df):
        self.df = self._prepare_data(df.copy())
    
    def _prepare_data(self, df):
        """Подготовка данных: преобразование дат и очистка"""
        date_columns = [
            'Принят в обработку', 'Дата отгрузки', 'Дата доставки',
            'Фактическая дата передачи в доставку'
        ]
        
        for col in date_columns:
            if col in df.columns:
                # Правильный парсинг дат с учетом формата dd.mm.yyyy
                df[col] = pd.to_datetime(df[col], format='%d.%m.%Y %H:%M', dayfirst=True, errors='coerce')
        
        # Преобразование числовых столбцов
        numeric_columns = ['Сумма отправления', 'Количество']
        for col in numeric_columns:
            if col in df.columns:
                # Удаляем запятые из чисел и преобразуем в float
                df[col] = df[col].astype(str).str.replace(',', '.').str.replace(' ', '')
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Удаление строк с некорректными датами
        df = df.dropna(subset=['Принят в обработку'])
        
        return df
    
    def calculate_delivery_metrics(self):
        """Расчет основных метрик доставки"""
        delivered_orders = self.df[self.df['Статус'] == 'Доставлен'].copy()
        
        if len(delivered_orders) == 0:
            return {
                'avg_delivery_time': 0,
                'median_delivery_time': 0,
                'avg_processing_time': 0,
                'avg_shipping_time': 0
            }
        
        # Время от обработки до доставки
        delivery_times = (delivered_orders['Дата доставки'] - 
                         delivered_orders['Принят в обработку']).dt.total_seconds() / (24 * 3600)
        
        # Время от обработки до передачи в доставку
        processing_times = (delivered_orders['Фактическая дата передачи в доставку'] - 
                           delivered_orders['Принят в обработку']).dt.total_seconds() / (24 * 3600)
        
        # Время от передачи в доставку до доставки
        shipping_times = (delivered_orders['Дата доставки'] - 
                         delivered_orders['Фактическая дата передачи в доставку']).dt.total_seconds() / (24 * 3600)
        
        return {
            'avg_delivery_time': delivery_times.mean() if not delivery_times.empty else 0,
            'median_delivery_time': delivery_times.median() if not delivery_times.empty else 0,
            'avg_processing_time': processing_times.mean() if not processing_times.empty else 0,
            'avg_shipping_time': shipping_times.mean() if not shipping_times.empty else 0
        }
    
    def get_daily_orders(self):
        """Получение количества заказов по дням"""
        daily_orders = self.df.groupby(self.df['Принят в обработку'].dt.date).size().reset_index()
        daily_orders.columns = ['date', 'count']
        return daily_orders
    
    def get_delivery_delays(self):
        """Анализ задержек доставки"""
        delivered_orders = self.df[self.df['Статус'] == 'Доставлен'].copy()
        
        if len(delivered_orders) == 0:
            return pd.DataFrame()
        
        # Предполагаем стандартное время доставки 3 дня
        standard_delivery_days = 3
        
        delivered_orders['expected_delivery'] = (delivered_orders['Принят в обработку'] + 
                                               timedelta(days=standard_delivery_days))
        
        delivered_orders['delay_days'] = (delivered_orders['Дата доставки'] - 
                                        delivered_orders['expected_delivery']).dt.total_seconds() / (24 * 3600)
        
        # Только задержки (положительные значения)
        delays = delivered_orders[delivered_orders['delay_days'] > 0][['Номер заказа', 'delay_days']]
        
        return delays
    
    def get_top_products_by_quantity(self, n=10):
        """Топ товаров по количеству"""
        return self.df.groupby('Наименование товара')['Количество'].sum().sort_values(ascending=False).head(n)
    
    def get_top_products_by_revenue(self, n=10):
        """Топ товаров по выручке (только доставленные заказы)"""
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        if len(delivered_df) == 0:
            return pd.Series(dtype=float)
        return delivered_df.groupby('Наименование товара')['Сумма отправления'].sum().sort_values(ascending=False).head(n)
    
    def get_status_distribution(self):
        """Распределение статусов заказов"""
        return self.df['Статус'].value_counts()
    
    def generate_pdf_report(self):
        """Генерация PDF отчета"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Регистрация шрифтов с поддержкой кириллицы
        registered_fonts = register_fonts()
        
        # Выбор шрифта для использования
        if 'DejaVuSans' in registered_fonts:
            font_name = 'DejaVuSans'
            font_bold = 'DejaVuSans-Bold'
        elif 'Arial' in registered_fonts:
            font_name = 'Arial'
            font_bold = 'Arial-Bold'
        elif 'HeiseiKakuGo-W5' in registered_fonts:
            font_name = 'HeiseiKakuGo-W5'
            font_bold = 'HeiseiKakuGo-W5'
        else:
            font_name = 'Helvetica'
            font_bold = 'Helvetica-Bold'
        
        styles = getSampleStyleSheet()
        story = []
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Центрирование
            fontName=font_bold
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            fontName=font_bold
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName=font_name
        )
        
        story.append(Paragraph("Отчет по анализу доставки заказов", title_style))
        story.append(Paragraph(f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y в %H:%M')}", normal_style))
        story.append(Spacer(1, 20))
        
        # Основные метрики
        story.append(Paragraph("Основные метрики доставки", heading_style))
        
        metrics = self.calculate_delivery_metrics()
        
        metrics_data = [
            ['Метрика', 'Значение'],
            ['Всего заказов', str(len(self.df))],
            ['Доставлено заказов', str(len(self.df[self.df['Статус'] == 'Доставлен']))],
            ['Среднее время доставки', f"{metrics['avg_delivery_time']:.1f} дней"],
            ['Медианное время доставки', f"{metrics['median_delivery_time']:.1f} дней"],
            ['Среднее время обработки', f"{metrics['avg_processing_time']:.1f} дней"],
            ['Среднее время в доставке', f"{metrics['avg_shipping_time']:.1f} дней"],
            ['Средняя сумма заказа', f"{self.df['Сумма отправления'].mean():.2f} ₽"],
            ['Общая выручка', f"{self.df['Сумма отправления'].sum():.2f} ₽"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Топ товары по количеству
        story.append(Paragraph("Топ-10 товаров по количеству", heading_style))
        
        top_quantity = self.get_top_products_by_quantity(10)
        quantity_data = [['Товар', 'Количество']]
        for product, qty in top_quantity.items():
            quantity_data.append([product[:50], str(qty)])  # Ограничиваем длину названия
        
        quantity_table = Table(quantity_data)
        quantity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(quantity_table)
        story.append(Spacer(1, 20))
        
        # Топ товары по выручке
        story.append(Paragraph("Топ-10 товаров по выручке", heading_style))
        
        top_revenue = self.get_top_products_by_revenue(10)
        revenue_data = [['Товар', 'Выручка (₽)']]
        for product, revenue in top_revenue.items():
            revenue_data.append([product[:50], f"{revenue:.2f}"])
        
        revenue_table = Table(revenue_data)
        revenue_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(revenue_table)
        story.append(Spacer(1, 20))
        
        # Распределение статусов
        story.append(Paragraph("Распределение статусов заказов", heading_style))
        
        status_dist = self.get_status_distribution()
        status_data = [['Статус', 'Количество', 'Процент']]
        total_orders = len(self.df)
        for status, count in status_dist.items():
            percentage = (count / total_orders) * 100
            status_data.append([status, str(count), f"{percentage:.1f}%"])
        
        status_table = Table(status_data)
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(status_table)
        story.append(Spacer(1, 20))
        
        # Анализ задержек
        story.append(Paragraph("Анализ задержек доставки", heading_style))
        
        delays = self.get_delivery_delays()
        if len(delays) > 0:
            delay_data = [['Метрика', 'Значение']]
            delay_data.append(['Заказов с задержкой', str(len(delays))])
            delay_data.append(['Средняя задержка (дни)', f"{delays['delay_days'].mean():.1f}"])
            delay_data.append(['Максимальная задержка (дни)', f"{delays['delay_days'].max():.1f}"])
            delay_percentage = (len(delays) / len(self.df[self.df['Статус'] == 'Доставлен'])) * 100 if len(self.df[self.df['Статус'] == 'Доставлен']) > 0 else 0
            delay_data.append(['Процент задержанных заказов', f"{delay_percentage:.1f}%"])
            
            delay_table = Table(delay_data)
            delay_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(delay_table)
        else:
            story.append(Paragraph("Задержек доставки не обнаружено", normal_style))
        
        # Сборка документа
        doc.build(story)
        buffer.seek(0)
        return buffer
    

    
    def generate_excel_report(self):
        """Генерация Excel отчета"""
        buffer = BytesIO()
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Стили для форматирования
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Лист 1: Основные метрики
        ws_metrics = wb.create_sheet("Основные метрики")
        
        # Заголовок отчета
        ws_metrics['A1'] = "Отчет по анализу доставки заказов"
        ws_metrics['A1'].font = title_font
        ws_metrics['A1'].alignment = center_alignment
        ws_metrics.merge_cells('A1:B1')
        
        # Дата генерации
        ws_metrics['A2'] = f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws_metrics.merge_cells('A2:B2')
        
        # Основные метрики
        metrics = self.calculate_delivery_metrics()
        
        metrics_data = [
            ['Метрика', 'Значение'],
            ['Всего заказов', len(self.df)],
            ['Доставлено заказов', len(self.df[self.df['Статус'] == 'Доставлен'])],
            ['Среднее время доставки', f"{metrics['avg_delivery_time']:.1f} дней"],
            ['Медианное время доставки', f"{metrics['median_delivery_time']:.1f} дней"],
            ['Среднее время обработки', f"{metrics['avg_processing_time']:.1f} дней"],
            ['Среднее время в доставке', f"{metrics['avg_shipping_time']:.1f} дней"],
            ['Средняя сумма заказа', f"{self.df['Сумма отправления'].mean():.2f} ₽"],
            ['Общая выручка', f"{self.df['Сумма отправления'].sum():.2f} ₽"]
        ]
        
        # Заполнение данных метрик
        start_row = 4
        for row_idx, row_data in enumerate(metrics_data):
            for col_idx, value in enumerate(row_data):
                cell = ws_metrics.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                cell.border = border
                if row_idx == 0:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for col_idx in range(1, 3):  # Колонки A и B
            max_length = 0
            column_letter = chr(64 + col_idx)  # A=65, B=66
            for row in ws_metrics.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws_metrics.column_dimensions[column_letter].width = adjusted_width
        
        # Лист 2: Топ товары по количеству
        ws_quantity = wb.create_sheet("Топ товары (количество)")
        
        ws_quantity['A1'] = "Топ-10 товаров по количеству"
        ws_quantity['A1'].font = title_font
        ws_quantity['A1'].alignment = center_alignment
        ws_quantity.merge_cells('A1:B1')
        
        top_quantity = self.get_top_products_by_quantity(10)
        quantity_data = [['Товар', 'Количество']]
        for product, qty in top_quantity.items():
            quantity_data.append([product, qty])
        
        # Заполнение данных
        start_row = 3
        for row_idx, row_data in enumerate(quantity_data):
            for col_idx, value in enumerate(row_data):
                cell = ws_quantity.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                cell.border = border
                if row_idx == 0:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for col_idx in range(1, 3):  # Колонки A и B
            max_length = 0
            column_letter = chr(64 + col_idx)  # A=65, B=66
            for row in ws_quantity.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 80)
            ws_quantity.column_dimensions[column_letter].width = adjusted_width
        
        # Лист 3: Топ товары по выручке
        ws_revenue = wb.create_sheet("Топ товары (выручка)")
        
        ws_revenue['A1'] = "Топ-10 товаров по выручке"
        ws_revenue['A1'].font = title_font
        ws_revenue['A1'].alignment = center_alignment
        ws_revenue.merge_cells('A1:B1')
        
        top_revenue = self.get_top_products_by_revenue(10)
        revenue_data = [['Товар', 'Выручка (₽)']]
        for product, revenue in top_revenue.items():
            revenue_data.append([product, f"{revenue:.2f}"])
        
        # Заполнение данных
        start_row = 3
        for row_idx, row_data in enumerate(revenue_data):
            for col_idx, value in enumerate(row_data):
                cell = ws_revenue.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                cell.border = border
                if row_idx == 0:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for col_idx in range(1, 3):  # Колонки A и B
            max_length = 0
            column_letter = chr(64 + col_idx)  # A=65, B=66
            for row in ws_revenue.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 80)
            ws_revenue.column_dimensions[column_letter].width = adjusted_width
        
        # Лист 4: Распределение статусов
        ws_status = wb.create_sheet("Статусы заказов")
        
        ws_status['A1'] = "Распределение статусов заказов"
        ws_status['A1'].font = title_font
        ws_status['A1'].alignment = center_alignment
        ws_status.merge_cells('A1:C1')
        
        status_dist = self.get_status_distribution()
        status_data = [['Статус', 'Количество', 'Процент']]
        total_orders = len(self.df)
        for status, count in status_dist.items():
            percentage = (count / total_orders) * 100
            status_data.append([status, count, f"{percentage:.1f}%"])
        
        # Заполнение данных
        start_row = 3
        for row_idx, row_data in enumerate(status_data):
            for col_idx, value in enumerate(row_data):
                cell = ws_status.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                cell.border = border
                if row_idx == 0:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for col_idx in range(1, 4):  # Колонки A, B, C
            max_length = 0
            column_letter = chr(64 + col_idx)  # A=65, B=66, C=67
            for row in ws_status.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 30)
            ws_status.column_dimensions[column_letter].width = adjusted_width
        
        # Лист 5: Анализ задержек
        ws_delays = wb.create_sheet("Анализ задержек")
        
        ws_delays['A1'] = "Анализ задержек доставки"
        ws_delays['A1'].font = title_font
        ws_delays['A1'].alignment = center_alignment
        ws_delays.merge_cells('A1:C1')
        
        delays = self.get_delivery_delays()
        if len(delays) > 0:
            ws_delays['A3'] = f"Количество заказов с задержкой: {len(delays)}"
            ws_delays['A4'] = f"Средняя задержка: {delays['delay_days'].mean():.1f} дней"
            ws_delays['A5'] = f"Максимальная задержка: {delays['delay_days'].max():.1f} дней"
            
            # Детальная таблица задержек
            delay_data = [['Номер заказа', 'Задержка (дней)']]
            for _, row in delays.head(20).iterrows():  # Показываем только первые 20
                delay_data.append([row['Номер заказа'], f"{row['delay_days']:.1f}"])
            
            # Заполнение данных
            start_row = 7
            for row_idx, row_data in enumerate(delay_data):
                for col_idx, value in enumerate(row_data):
                    cell = ws_delays.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                    cell.border = border
                    if row_idx == 0:  # Заголовок
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
        else:
            ws_delays['A3'] = "Задержек доставки не обнаружено"
        
        # Автоширина колонок
        for col_idx in range(1, 4):  # Колонки A, B, C
            max_length = 0
            column_letter = chr(64 + col_idx)  # A=65, B=66, C=67
            for row in ws_delays.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 40)
            ws_delays.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в буфер
        wb.save(buffer)
        buffer.seek(0)
        return buffer