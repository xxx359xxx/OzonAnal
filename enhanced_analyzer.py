import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from io import BytesIO
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from utils import register_fonts

class EnhancedOrderAnalyzer:
    """
    Расширенный анализатор заказов для работы с новым форматом CSV (25 колонок)
    Поддерживает анализ скидок, акций, веса товаров и других расширенных метрик
    """
    
    def __init__(self, df):
        self.original_df = df.copy()
        self.df = self._prepare_data(df)
        
    def _prepare_data(self, df):
        """Подготовка и очистка данных"""
        df_clean = df.copy()
        
        # Определяем колонки дат в новом формате
        date_columns = [
            'Принят в обработку',
            'Фактическая дата передачи в доставку',
            'Дата доставки'
        ]
        
        # Конвертация дат (формат YYYY-MM-DD HH:MM:SS)
        for col in date_columns:
            if col in df_clean.columns:
                df_clean[col] = pd.to_datetime(df_clean[col], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        
        # Конвертация числовых полей
        numeric_columns = [
            'Ваша цена',
            'Цена товара до скидок',
            'Скидка руб',
            'Скидка %',
            'Объемный вес товаров, кг',
            'Количество',
            'Сумма отправления'
        ]
        
        for col in numeric_columns:
            if col in df_clean.columns:
                # Заменяем запятые на точки и конвертируем в числа
                df_clean[col] = df_clean[col].astype(str).str.replace(',', '.', regex=False)
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
        
        return df_clean
    
    def get_basic_metrics(self):
        """Базовые метрики заказов"""
        total_orders = len(self.df)
        delivered_orders = len(self.df[self.df['Статус'] == 'Доставлен'])
        cancelled_orders = len(self.df[self.df['Статус'] == 'Отменён'])
        in_delivery = len(self.df[self.df['Статус'] == 'Доставляется'])
        
        # Финансовые метрики - считаем выручку только по доставленным заказам
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) > 0:
            total_revenue = delivered_df['Сумма отправления'].sum() if 'Сумма отправления' in delivered_df.columns else delivered_df['Ваша цена'].sum()
            avg_order_value = delivered_df['Сумма отправления'].mean() if 'Сумма отправления' in delivered_df.columns else delivered_df['Ваша цена'].mean()
        else:
            total_revenue = 0
            avg_order_value = 0
        total_discount_amount = self.df['Скидка руб'].sum()
        
        return {
            'total_orders': total_orders,
            'delivered_orders': delivered_orders,
            'cancelled_orders': cancelled_orders,
            'in_delivery': in_delivery,
            'delivery_rate': (delivered_orders / total_orders * 100) if total_orders > 0 else 0,
            'cancellation_rate': (cancelled_orders / total_orders * 100) if total_orders > 0 else 0,
            'total_revenue': total_revenue,
            'avg_order_value': avg_order_value,
            'total_discount_amount': total_discount_amount
        }
    
    def analyze_discounts(self):
        """Анализ скидок и акций"""
        # Фильтруем заказы со скидками
        discounted_orders = self.df[self.df['Скидка руб'] > 0]
        
        if len(discounted_orders) == 0:
            return {
                'orders_with_discount': 0,
                'discount_percentage': 0,
                'avg_discount_amount': 0,
                'avg_discount_percent': 0,
                'total_savings': 0
            }
        
        total_savings = discounted_orders['Скидка руб'].sum()
        avg_discount_amount = discounted_orders['Скидка руб'].mean()
        avg_discount_percent = discounted_orders['Скидка %'].mean()
        
        return {
            'orders_with_discount': len(discounted_orders),
            'discount_percentage': (len(discounted_orders) / len(self.df) * 100),
            'avg_discount_amount': avg_discount_amount,
            'avg_discount_percent': avg_discount_percent,
            'total_savings': total_savings,
            'max_discount_amount': discounted_orders['Скидка руб'].max(),
            'max_discount_percent': discounted_orders['Скидка %'].max()
        }
    
    def analyze_product_categories(self):
        """Анализ категорий товаров"""
        # Используем только доставленные заказы для анализа товаров
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) == 0:
            return pd.DataFrame()
        
        # Группировка по товарам
        revenue_column = 'Сумма отправления' if 'Сумма отправления' in delivered_df.columns else 'Ваша цена'
        agg_dict = {
            'Количество': 'sum',
            revenue_column: ['sum', 'mean'],
            'Скидка руб': 'sum'
        }
        
        # Добавляем вес только если колонка существует
        if 'Объемный вес товаров, кг' in delivered_df.columns:
            agg_dict['Объемный вес товаров, кг'] = 'mean'
        
        product_stats = delivered_df.groupby('Наименование товара').agg(agg_dict).round(2)
        
        # Упрощаем названия колонок
        if 'Объемный вес товаров, кг' in self.df.columns:
            product_stats.columns = [
                'total_quantity',
                'total_revenue',
                'avg_price',
                'total_discount',
                'avg_weight'
            ]
        else:
            product_stats.columns = [
                'total_quantity',
                'total_revenue',
                'avg_price',
                'total_discount'
            ]
            # Добавляем колонку avg_weight с NaN для совместимости
            product_stats['avg_weight'] = pd.NA
        
        # Сортируем по выручке
        product_stats = product_stats.sort_values('total_revenue', ascending=False)
        
        return product_stats
    
    def analyze_product_by_sku(self):
        """Анализ товаров по артикулам (SKU)"""
        # Используем только доставленные заказы
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) == 0 or 'Артикул' not in delivered_df.columns:
            return pd.DataFrame()
        
        # Группировка по артикулам
        revenue_column = 'Сумма отправления' if 'Сумма отправления' in delivered_df.columns else 'Ваша цена'
        agg_dict = {
            'Количество': 'sum',
            revenue_column: ['sum', 'mean'],
            'Скидка руб': 'sum',
            'Наименование товара': 'first'  # Берем первое название товара для артикула
        }
        
        # Добавляем вес если есть
        if 'Объемный вес товаров, кг' in delivered_df.columns:
            agg_dict['Объемный вес товаров, кг'] = 'mean'
        
        sku_stats = delivered_df.groupby('Артикул').agg(agg_dict).round(2)
        
        # Упрощаем названия колонок
        if 'Объемный вес товаров, кг' in self.df.columns:
            sku_stats.columns = [
                'total_quantity',
                'total_revenue', 
                'avg_price',
                'total_discount',
                'product_name',
                'avg_weight'
            ]
        else:
            sku_stats.columns = [
                'total_quantity',
                'total_revenue',
                'avg_price', 
                'total_discount',
                'product_name'
            ]
            sku_stats['avg_weight'] = pd.NA
        
        # Добавляем расчетные метрики
        sku_stats['discount_rate'] = (sku_stats['total_discount'] / (sku_stats['total_revenue'] + sku_stats['total_discount']) * 100).round(2)
        sku_stats['revenue_per_unit'] = (sku_stats['total_revenue'] / sku_stats['total_quantity']).round(2)
        
        # Сортируем по выручке
        sku_stats = sku_stats.sort_values('total_revenue', ascending=False)
        
        return sku_stats
    
    def get_sku_abc_analysis(self):
        """ABC анализ артикулов по выручке"""
        sku_stats = self.analyze_product_by_sku()
        
        if len(sku_stats) == 0:
            return {
                'category_A': [],
                'category_B': [],
                'category_C': [],
                'revenue_share_A': 0,
                'revenue_share_B': 0,
                'revenue_share_C': 0,
                'cumulative_A': 0,
                'cumulative_B': 0,
                'cumulative_C': 100
            }
        
        # Рассчитываем кумулятивную долю выручки
        total_revenue = sku_stats['total_revenue'].sum()
        sku_stats['revenue_share'] = (sku_stats['total_revenue'] / total_revenue * 100).round(2)
        sku_stats['cumulative_share'] = sku_stats['revenue_share'].cumsum().round(2)
        
        # Классификация ABC
        def classify_abc(cumulative_share):
            if cumulative_share <= 80:
                return 'A'
            elif cumulative_share <= 95:
                return 'B'
            else:
                return 'C'
        
        sku_stats['abc_category'] = sku_stats['cumulative_share'].apply(classify_abc)
        
        # Разделяем по категориям
        category_A = sku_stats[sku_stats['abc_category'] == 'A'].index.tolist()
        category_B = sku_stats[sku_stats['abc_category'] == 'B'].index.tolist()
        category_C = sku_stats[sku_stats['abc_category'] == 'C'].index.tolist()
        
        # Рассчитываем доли выручки
        revenue_share_A = sku_stats[sku_stats['abc_category'] == 'A']['revenue_share'].sum()
        revenue_share_B = sku_stats[sku_stats['abc_category'] == 'B']['revenue_share'].sum()
        revenue_share_C = sku_stats[sku_stats['abc_category'] == 'C']['revenue_share'].sum()
        
        # Кумулятивные доли
        cumulative_A = revenue_share_A
        cumulative_B = revenue_share_A + revenue_share_B
        cumulative_C = 100.0
        
        return {
            'category_A': category_A,
            'category_B': category_B,
            'category_C': category_C,
            'revenue_share_A': revenue_share_A,
            'revenue_share_B': revenue_share_B,
            'revenue_share_C': revenue_share_C,
            'cumulative_A': cumulative_A,
            'cumulative_B': cumulative_B,
            'cumulative_C': cumulative_C
        }
    
    def get_sku_performance_metrics(self):
        """Метрики эффективности артикулов"""
        sku_stats = self.analyze_product_by_sku()
        
        if len(sku_stats) == 0:
            return {
                'total_skus': 0,
                'total_unique_skus': 0,
                'avg_revenue_per_sku': 0,
                'median_revenue_per_sku': 0,
                'avg_quantity_per_sku': 0,
                'top_sku_by_revenue': 'N/A',
                'top_sku_revenue': 0,
                'top_sku_by_quantity': 'N/A',
                'top_sku_quantity': 0,
                'a_category_skus': 0,
                'b_category_skus': 0,
                'c_category_skus': 0,
                'a_category_revenue_share': 0,
                'top_10_percent_revenue_share': 0,
                'gini_coefficient': 0,
                'high_revenue_skus': 0,
                'high_revenue_share': 0
            }
        
        total_skus = len(sku_stats)
        total_revenue = sku_stats['total_revenue'].sum()
        total_quantity = sku_stats['total_quantity'].sum()
        
        # Топ артикулы
        top_sku_by_revenue = sku_stats.head(1).index[0] if len(sku_stats) > 0 else 'N/A'
        top_sku_revenue = sku_stats.head(1)['total_revenue'].iloc[0] if len(sku_stats) > 0 else 0
        
        top_sku_by_quantity = sku_stats.nlargest(1, 'total_quantity').index[0] if len(sku_stats) > 0 else 'N/A'
        top_sku_quantity = sku_stats.nlargest(1, 'total_quantity')['total_quantity'].iloc[0] if len(sku_stats) > 0 else 0
        
        # ABC анализ
        abc_analysis = self.get_sku_abc_analysis()
        a_skus = len(abc_analysis['category_A'])
        b_skus = len(abc_analysis['category_B'])
        c_skus = len(abc_analysis['category_C'])
        a_revenue_share = abc_analysis['revenue_share_A']
        
        # Топ-10% артикулов по выручке
        top_10_percent_count = max(1, int(total_skus * 0.1))
        top_10_percent_revenue = sku_stats.head(top_10_percent_count)['total_revenue'].sum()
        top_10_percent_revenue_share = (top_10_percent_revenue / total_revenue * 100) if total_revenue > 0 else 0
        
        # Коэффициент Джини (концентрация выручки)
        revenues = sku_stats['total_revenue'].values
        revenues_sorted = np.sort(revenues)
        n = len(revenues_sorted)
        cumsum = np.cumsum(revenues_sorted)
        gini_coefficient = (n + 1 - 2 * np.sum(cumsum) / cumsum[-1]) / n if cumsum[-1] > 0 else 0
        
        # Высокодоходные артикулы (>10k₽)
        high_revenue_threshold = 10000
        high_revenue_mask = sku_stats['total_revenue'] > high_revenue_threshold
        high_revenue_skus = high_revenue_mask.sum()
        high_revenue_total = sku_stats[high_revenue_mask]['total_revenue'].sum()
        high_revenue_share = (high_revenue_total / total_revenue * 100) if total_revenue > 0 else 0
        
        return {
            'total_skus': total_skus,
            'total_unique_skus': total_skus,  # Добавляем для совместимости с отчетом
            'avg_revenue_per_sku': (total_revenue / total_skus).round(2) if total_skus > 0 else 0,
            'median_revenue_per_sku': sku_stats['total_revenue'].median() if len(sku_stats) > 0 else 0,
            'avg_quantity_per_sku': (total_quantity / total_skus).round(2) if total_skus > 0 else 0,
            'top_sku_by_revenue': top_sku_by_revenue,
            'top_sku_revenue': top_sku_revenue,
            'top_sku_by_quantity': top_sku_by_quantity,
            'top_sku_quantity': top_sku_quantity,
            'a_category_skus': a_skus,
            'b_category_skus': b_skus,
            'c_category_skus': c_skus,
            'a_category_revenue_share': a_revenue_share,
            'top_10_percent_revenue_share': round(top_10_percent_revenue_share, 1),
            'gini_coefficient': round(gini_coefficient, 3),
            'high_revenue_skus': high_revenue_skus,
            'high_revenue_share': round(high_revenue_share, 1)
        }
    
    def analyze_delivery_performance(self):
        """Анализ производительности доставки"""
        delivered_orders = self.df[self.df['Статус'] == 'Доставлен'].copy()
        
        if len(delivered_orders) == 0:
            return {
                'avg_delivery_time': 0,
                'median_delivery_time': 0,
                'delivery_time_std': 0,
                'on_time_delivery_rate': 0
            }
        
        # Вычисляем время доставки
        delivered_orders['delivery_time'] = (
            delivered_orders['Дата доставки'] - 
            delivered_orders['Принят в обработку']
        ).dt.days
        
        # Убираем отрицательные значения и выбросы
        delivered_orders = delivered_orders[delivered_orders['delivery_time'] >= 0]
        delivered_orders = delivered_orders[delivered_orders['delivery_time'] <= 30]  # Максимум 30 дней
        
        if len(delivered_orders) == 0:
            return {
                'avg_delivery_time': 0,
                'median_delivery_time': 0,
                'delivery_time_std': 0,
                'on_time_delivery_rate': 0
            }
        
        avg_delivery_time = delivered_orders['delivery_time'].mean()
        median_delivery_time = delivered_orders['delivery_time'].median()
        delivery_time_std = delivered_orders['delivery_time'].std()
        
        # Считаем доставку вовремя, если <= 5 дней
        on_time_orders = len(delivered_orders[delivered_orders['delivery_time'] <= 5])
        on_time_rate = (on_time_orders / len(delivered_orders) * 100) if len(delivered_orders) > 0 else 0
        
        return {
            'avg_delivery_time': avg_delivery_time,
            'median_delivery_time': median_delivery_time,
            'delivery_time_std': delivery_time_std,
            'on_time_delivery_rate': on_time_rate,
            'total_delivered': len(delivered_orders)
        }
    
    def analyze_weight_logistics(self):
        """Анализ логистики по весу товаров"""
        # Проверяем наличие колонки веса
        if 'Объемный вес товаров, кг' not in self.df.columns:
            return {
                'total_weight_kg': 0,
                'avg_weight_per_order': 0,
                'heavy_orders_count': 0,
                'light_orders_count': 0,
                'heavy_orders_percentage': 0,
                'light_orders_percentage': 0,
                'weight_data_available': False
            }
        
        weight_data = self.df[self.df['Объемный вес товаров, кг'].notna()]
        
        if len(weight_data) == 0:
            return {
                'total_weight_kg': 0,
                'avg_weight_per_order': 0,
                'heavy_orders_count': 0,
                'light_orders_count': 0,
                'heavy_orders_percentage': 0,
                'light_orders_percentage': 0,
                'weight_data_available': False
            }
        
        total_weight_kg = weight_data['Объемный вес товаров, кг'].sum()
        avg_weight_per_order = weight_data['Объемный вес товаров, кг'].mean()
        
        # Классификация по весу (тяжелые > 2кг, легкие <= 0.5кг)
        heavy_orders = len(weight_data[weight_data['Объемный вес товаров, кг'] > 2.0])
        light_orders = len(weight_data[weight_data['Объемный вес товаров, кг'] <= 0.5])
        
        return {
            'total_weight_kg': total_weight_kg,
            'avg_weight_per_order': avg_weight_per_order,
            'heavy_orders_count': heavy_orders,
            'light_orders_count': light_orders,
            'heavy_orders_percentage': (heavy_orders / len(weight_data) * 100),
            'light_orders_percentage': (light_orders / len(weight_data) * 100),
            'weight_data_available': True
        }
    
    def get_regional_analysis(self):
        """Анализ по регионам доставки"""
        if 'Регион доставки' not in self.df.columns:
            return {}
        
        # Используем только доставленные заказы
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) == 0:
            return {}
        
        revenue_column = 'Сумма отправления' if 'Сумма отправления' in delivered_df.columns else 'Ваша цена'
        regional_stats = delivered_df.groupby('Регион доставки').agg({
            'Номер заказа': 'count',
            revenue_column: ['sum', 'mean'],
            'Скидка руб': 'sum'
        }).round(2)
        
        regional_stats.columns = ['order_count', 'total_revenue', 'avg_order_value', 'total_discount']
        regional_stats = regional_stats.sort_values('total_revenue', ascending=False)
        
        return regional_stats
    
    def get_time_series_analysis(self):
        """Временной анализ заказов"""
        if 'Принят в обработку' not in self.df.columns:
            return pd.DataFrame()
        
        # Используем только доставленные заказы
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) == 0:
            return pd.DataFrame()
        
        # Группируем по дням
        revenue_column = 'Сумма отправления' if 'Сумма отправления' in delivered_df.columns else 'Ваша цена'
        daily_stats = delivered_df.groupby(delivered_df['Принят в обработку'].dt.date).agg({
            'Номер заказа': 'count',
            revenue_column: 'sum',
            'Скидка руб': 'sum',
            'Количество': 'sum'
        }).round(2)
        
        daily_stats.columns = ['orders_count', 'daily_revenue', 'daily_discount', 'items_sold']
        daily_stats.index.name = 'date'
        
        return daily_stats.reset_index()
    
    def get_monthly_analysis(self):
        """Месячный анализ для выявления трендов и успешных периодов"""
        if 'Принят в обработку' not in self.df.columns:
            return pd.DataFrame()
        
        # Используем только доставленные заказы
        delivered_df = self.df[self.df['Статус'] == 'Доставлен']
        
        if len(delivered_df) == 0:
            return pd.DataFrame()
        
        # Добавляем колонку месяца
        df_with_month = delivered_df.copy()
        df_with_month['month_year'] = df_with_month['Принят в обработку'].dt.to_period('M')
        
        # Группируем по месяцам
        revenue_column = 'Сумма отправления' if 'Сумма отправления' in delivered_df.columns else 'Ваша цена'
        agg_dict = {
            'Номер заказа': 'count',
            revenue_column: ['sum', 'mean'],
            'Скидка руб': 'sum',
            'Количество': 'sum'
        }
        
        # Добавляем вес только если колонка существует
        if 'Объемный вес товаров, кг' in df_with_month.columns:
            agg_dict['Объемный вес товаров, кг'] = ['mean', 'sum']
        
        monthly_stats = df_with_month.groupby('month_year').agg(agg_dict).round(2)
        
        # Упрощаем названия колонок
        if 'Объемный вес товаров, кг' in df_with_month.columns:
            monthly_stats.columns = [
                'orders_count', 'total_revenue', 'avg_order_value', 
                'total_discount', 'total_quantity', 'avg_weight_kg', 'total_weight_kg'
            ]
        else:
            monthly_stats.columns = [
                'orders_count', 'total_revenue', 'avg_order_value', 
                'total_discount', 'total_quantity'
            ]
            # Добавляем колонки веса с NaN для совместимости
            monthly_stats['avg_weight_kg'] = pd.NA
            monthly_stats['total_weight_kg'] = pd.NA
        
        # Добавляем расчетные метрики
        monthly_stats['revenue_per_order'] = monthly_stats['total_revenue'] / monthly_stats['orders_count']
        monthly_stats['discount_rate'] = (monthly_stats['total_discount'] / monthly_stats['total_revenue'] * 100).fillna(0)
        monthly_stats['avg_items_per_order'] = monthly_stats['total_quantity'] / monthly_stats['orders_count']
        
        # Нормализуем метрики для расчета рейтинга
        revenue_norm = (monthly_stats['total_revenue'] - monthly_stats['total_revenue'].min()) / (monthly_stats['total_revenue'].max() - monthly_stats['total_revenue'].min() + 1e-10)
        volume_norm = (monthly_stats['orders_count'] * monthly_stats['avg_order_value'] - (monthly_stats['orders_count'] * monthly_stats['avg_order_value']).min()) / ((monthly_stats['orders_count'] * monthly_stats['avg_order_value']).max() - (monthly_stats['orders_count'] * monthly_stats['avg_order_value']).min() + 1e-10)
        discount_eff_norm = (100 - monthly_stats['discount_rate']) / 100
        turnover_norm = (monthly_stats['avg_items_per_order'] * monthly_stats['orders_count'] - (monthly_stats['avg_items_per_order'] * monthly_stats['orders_count']).min()) / ((monthly_stats['avg_items_per_order'] * monthly_stats['orders_count']).max() - (monthly_stats['avg_items_per_order'] * monthly_stats['orders_count']).min() + 1e-10)
        
        # Ранжируем месяцы по успешности (комплексная метрика)
        monthly_stats['success_rating'] = (
            revenue_norm * 40 +  # 40% - выручка
            volume_norm * 30 +  # 30% - объем * средний чек
            discount_eff_norm * 20 +  # 20% - эффективность скидок
            turnover_norm * 10  # 10% - товарооборот
        )
        
        # Добавляем колонку с названием месяца для отображения
        monthly_stats_reset = monthly_stats.reset_index()
        monthly_stats_reset['month'] = monthly_stats_reset['month_year'].astype(str)
        
        # Сортируем по дате для правильного отображения временных трендов
        monthly_stats_reset = monthly_stats_reset.sort_values('month_year')
        
        return monthly_stats_reset
    
    def generate_enhanced_excel_report(self):
        """Генерация расширенного Excel отчета"""
        buffer = BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Лист 1: Основные метрики
        ws_main = wb.create_sheet("Основные метрики")
        ws_main['A1'] = "Расширенный отчет по анализу заказов Ozon"
        ws_main['A1'].font = title_font
        ws_main.merge_cells('A1:B1')
        
        ws_main['A2'] = f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws_main.merge_cells('A2:B2')
        
        # Основные метрики
        basic_metrics = self.get_basic_metrics()
        
        metrics_data = [
            ['Метрика', 'Значение'],
            ['Всего заказов', basic_metrics['total_orders']],
            ['Доставлено заказов', basic_metrics['delivered_orders']],
            ['Отменено заказов', basic_metrics['cancelled_orders']],
            ['В доставке', basic_metrics['in_delivery']],
            ['Процент доставки', f"{basic_metrics['delivery_rate']:.1f}%"],
            ['Процент отмен', f"{basic_metrics['cancellation_rate']:.1f}%"],
            ['Общая выручка', f"{basic_metrics['total_revenue']:.2f} ₽"],
            ['Средний чек', f"{basic_metrics['avg_order_value']:.2f} ₽"],
            ['Общая сумма скидок', f"{basic_metrics['total_discount_amount']:.2f} ₽"]
        ]
        
        self._fill_worksheet_data(ws_main, metrics_data, 4, header_font, header_fill, border, center_alignment)
        
        # Лист 2: Товарная аналитика по артикулам
        ws_sku = wb.create_sheet("Анализ артикулов")
        ws_sku['A1'] = "Товарная аналитика по артикулам"
        ws_sku['A1'].font = title_font
        
        # Анализ товаров по артикулам
        sku_df = self.analyze_product_by_sku()
        sku_data = [['Артикул', 'Товар', 'Количество', 'Выручка (₽)', 'Средняя цена (₽)', 'Скидка (₽)']]
        
        for _, row in sku_df.head(20).iterrows():
            sku_data.append([
                row.name,  # Артикул (индекс)
                row['product_name'][:30] + '...' if len(str(row['product_name'])) > 30 else row['product_name'],
                int(row['total_quantity']),
                f"{row['total_revenue']:.2f}",
                f"{row['avg_price']:.2f}",
                f"{row['total_discount']:.2f}"
            ])
        
        self._fill_worksheet_data(ws_sku, sku_data, 3, header_font, header_fill, border, center_alignment)
        
        # ABC-анализ артикулов
        abc_analysis = self.get_sku_abc_analysis()
        ws_sku['A26'] = "ABC-анализ артикулов"
        ws_sku['A26'].font = Font(bold=True, size=14)
        
        abc_data = [
            ['Категория', 'Количество артикулов', 'Доля выручки (%)', 'Накопленная доля (%)'],
            ['A (топ)', len(abc_analysis['category_A']), f"{abc_analysis['revenue_share_A']:.1f}", f"{abc_analysis['cumulative_A']:.1f}"],
            ['B (средние)', len(abc_analysis['category_B']), f"{abc_analysis['revenue_share_B']:.1f}", f"{abc_analysis['cumulative_B']:.1f}"],
            ['C (низкие)', len(abc_analysis['category_C']), f"{abc_analysis['revenue_share_C']:.1f}", "100.0"]
        ]
        
        self._fill_worksheet_data(ws_sku, abc_data, 28, header_font, header_fill, border, center_alignment)
        
        # Метрики эффективности артикулов
        performance = self.get_sku_performance_metrics()
        ws_sku['A34'] = "Ключевые метрики артикулов"
        ws_sku['A34'].font = Font(bold=True, size=14)
        
        perf_data = [
            ['Метрика', 'Значение'],
            ['Всего уникальных артикулов', performance['total_unique_skus']],
            ['Средняя выручка на артикул', f"{performance['avg_revenue_per_sku']:.2f} ₽"],
            ['Медианная выручка на артикул', f"{performance['median_revenue_per_sku']:.2f} ₽"],
            ['Топ-10% артикулов дают выручки', f"{performance['top_10_percent_revenue_share']:.1f}%"],
            ['Коэффициент концентрации (Джини)', f"{performance['gini_coefficient']:.3f}"],
            ['Артикулов с выручкой > 10000₽', performance['high_revenue_skus']],
            ['Доля высокодоходных артикулов', f"{performance['high_revenue_share']:.1f}%"]
        ]
        
        self._fill_worksheet_data(ws_sku, perf_data, 36, header_font, header_fill, border, center_alignment)

        # Лист 3: Анализ скидок
        ws_discounts = wb.create_sheet("Анализ скидок")
        ws_discounts['A1'] = "Анализ скидок и акций"
        ws_discounts['A1'].font = title_font

        discount_analysis = self.analyze_discounts()
        discount_data = [
            ['Метрика', 'Значение'],
            ['Заказов со скидкой', discount_analysis['orders_with_discount']],
            ['Процент заказов со скидкой', f"{discount_analysis['discount_percentage']:.1f}%"],
            ['Средняя скидка (руб)', f"{discount_analysis['avg_discount_amount']:.2f} ₽"],
            ['Средняя скидка (%)', f"{discount_analysis['avg_discount_percent']:.1f}%"],
            ['Общая экономия клиентов', f"{discount_analysis['total_savings']:.2f} ₽"],
            ['Максимальная скидка (руб)', f"{discount_analysis['max_discount_amount']:.2f} ₽"],
            ['Максимальная скидка (%)', f"{discount_analysis['max_discount_percent']:.1f}%"]
        ]
        
        self._fill_worksheet_data(ws_discounts, discount_data, 3, header_font, header_fill, border, center_alignment)
        
        # Лист 3: Анализ товаров
        ws_products = wb.create_sheet("Анализ товаров")
        ws_products['A1'] = "Топ-20 товаров по выручке"
        ws_products['A1'].font = title_font
        
        product_stats = self.analyze_product_categories().head(20)
        
        # Проверяем наличие данных о весе
        has_weight_data = 'Объемный вес товаров, кг' in self.df.columns
        
        if has_weight_data:
            product_data = [['Товар', 'Количество', 'Выручка (₽)', 'Средняя цена (₽)', 'Скидка (₽)', 'Средний вес (г)']]
            for product, row in product_stats.iterrows():
                product_data.append([
                    product[:50],  # Ограничиваем длину названия
                    int(row['total_quantity']),
                    f"{row['total_revenue']:.2f}",
                    f"{row['avg_price']:.2f}",
                    f"{row['total_discount']:.2f}",
                    f"{row['avg_weight']:.0f}" if pd.notna(row['avg_weight']) else "N/A"
                ])
        else:
            product_data = [['Товар', 'Количество', 'Выручка (₽)', 'Средняя цена (₽)', 'Скидка (₽)']]
            for product, row in product_stats.iterrows():
                product_data.append([
                    product[:50],  # Ограничиваем длину названия
                    int(row['total_quantity']),
                    f"{row['total_revenue']:.2f}",
                    f"{row['avg_price']:.2f}",
                    f"{row['total_discount']:.2f}"
                ])
        
        self._fill_worksheet_data(ws_products, product_data, 3, header_font, header_fill, border, center_alignment)
        
        # Лист 4: Анализ доставки
        ws_delivery = wb.create_sheet("Анализ доставки")
        ws_delivery['A1'] = "Производительность доставки"
        ws_delivery['A1'].font = title_font
        
        delivery_analysis = self.analyze_delivery_performance()
        delivery_data = [
            ['Метрика', 'Значение'],
            ['Среднее время доставки', f"{delivery_analysis['avg_delivery_time']:.1f} дней"],
            ['Медианное время доставки', f"{delivery_analysis['median_delivery_time']:.1f} дней"],
            ['Стандартное отклонение', f"{delivery_analysis['delivery_time_std']:.1f} дней"],
            ['Процент доставки вовремя (≤5 дней)', f"{delivery_analysis['on_time_delivery_rate']:.1f}%"],
            ['Всего доставленных заказов', delivery_analysis['total_delivered']]
        ]
        
        self._fill_worksheet_data(ws_delivery, delivery_data, 3, header_font, header_fill, border, center_alignment)
        
        # Лист 5: Анализ веса
        ws_weight = wb.create_sheet("Логистика по весу")
        ws_weight['A1'] = "Анализ веса товаров"
        ws_weight['A1'].font = title_font
        
        weight_analysis = self.analyze_weight_logistics()
        
        if weight_analysis.get('weight_data_available', False):
            weight_data = [
                ['Метрика', 'Значение'],
                ['Общий вес (кг)', f"{weight_analysis['total_weight_kg']:.2f}"],
                ['Средний вес заказа (г)', f"{weight_analysis['avg_weight_per_order']:.0f}"],
                ['Тяжелых заказов (>2кг)', weight_analysis['heavy_orders_count']],
                ['Легких заказов (≤500г)', weight_analysis['light_orders_count']],
                ['Процент тяжелых заказов', f"{weight_analysis['heavy_orders_percentage']:.1f}%"],
                ['Процент легких заказов', f"{weight_analysis['light_orders_percentage']:.1f}%"]
            ]
        else:
            weight_data = [
                ['Метрика', 'Значение'],
                ['Статус данных', 'Данные о весе товаров отсутствуют'],
                ['Общий вес (кг)', 'N/A'],
                ['Средний вес заказа (г)', 'N/A'],
                ['Тяжелых заказов (>2кг)', 'N/A'],
                ['Легких заказов (≤500г)', 'N/A'],
                ['Процент тяжелых заказов', 'N/A'],
                ['Процент легких заказов', 'N/A']
            ]
        
        self._fill_worksheet_data(ws_weight, weight_data, 3, header_font, header_fill, border, center_alignment)
        
        # Лист 6: Месячный анализ
        ws_monthly = wb.create_sheet("Месячный анализ")
        ws_monthly['A1'] = "Анализ по месяцам (ранжировано по успешности)"
        ws_monthly['A1'].font = title_font
        
        monthly_analysis = self.get_monthly_analysis()
        if not monthly_analysis.empty:
            # Проверяем наличие данных о весе
            has_weight_data = 'Объемный вес товаров, кг' in self.df.columns
            
            if has_weight_data:
                ws_monthly.merge_cells('A1:K1')
                monthly_data = [[
                    'Месяц', 'Заказов', 'Выручка (₽)', 'Средний чек (₽)', 
                    'Скидки (₽)', 'Товаров', 'Средний вес (г)', 'Общий вес (кг)',
                    'Выручка/заказ', 'Скидка %', 'Рейтинг успешности'
                ]]
                
                for _, row in monthly_analysis.iterrows():
                    monthly_data.append([
                        row['month'],
                        int(row['orders_count']),
                        f"{row['total_revenue']:.2f}",
                        f"{row['avg_order_value']:.2f}",
                        f"{row['total_discount']:.2f}",
                        int(row['total_quantity']),
                        f"{row['avg_weight_kg']:.0f}" if pd.notna(row['avg_weight_kg']) else "N/A",
                        f"{row['total_weight_kg']:.2f}" if pd.notna(row['total_weight_kg']) else "N/A",
                        f"{row['revenue_per_order']:.2f}",
                        f"{row['discount_rate']:.1f}%",
                        f"{row['success_rating']:.0f}"
                    ])
            else:
                ws_monthly.merge_cells('A1:I1')
                monthly_data = [[
                    'Месяц', 'Заказов', 'Выручка (₽)', 'Средний чек (₽)', 
                    'Скидки (₽)', 'Товаров', 'Выручка/заказ', 'Скидка %', 'Рейтинг успешности'
                ]]
                
                for _, row in monthly_analysis.iterrows():
                    monthly_data.append([
                        row['month'],
                        int(row['orders_count']),
                        f"{row['total_revenue']:.2f}",
                        f"{row['avg_order_value']:.2f}",
                        f"{row['total_discount']:.2f}",
                        int(row['total_quantity']),
                        f"{row['revenue_per_order']:.2f}",
                        f"{row['discount_rate']:.1f}%",
                        f"{row['success_rating']:.0f}"
                    ])
            
            self._fill_worksheet_data(ws_monthly, monthly_data, 3, header_font, header_fill, border, center_alignment)
            
            # Добавляем пояснение к рейтингу
            ws_monthly[f'A{len(monthly_data) + 5}'] = "Рейтинг успешности рассчитывается на основе:"
            ws_monthly[f'A{len(monthly_data) + 6}'] = "• 40% - общая выручка"
            ws_monthly[f'A{len(monthly_data) + 7}'] = "• 30% - объем заказов × средний чек"
            ws_monthly[f'A{len(monthly_data) + 8}'] = "• 20% - эффективность скидок"
            ws_monthly[f'A{len(monthly_data) + 9}'] = "• 10% - товарооборот"
        
        # Автоширина колонок для всех листов
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                if column_letter:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _fill_worksheet_data(self, ws, data, start_row, header_font, header_fill, border, center_alignment):
        """Вспомогательный метод для заполнения данных в лист Excel"""
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                cell.border = border
                if row_idx == 0:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
    
    def get_summary_insights(self):
        """Получение ключевых инсайтов для отчета"""
        basic_metrics = self.get_basic_metrics()
        discount_analysis = self.analyze_discounts()
        delivery_analysis = self.analyze_delivery_performance()
        weight_analysis = self.analyze_weight_logistics()
        
        insights = []
        
        # Инсайты по доставке
        if delivery_analysis['on_time_delivery_rate'] > 80:
            insights.append(f"✅ Отличная производительность доставки: {delivery_analysis['on_time_delivery_rate']:.1f}% заказов доставлено вовремя")
        elif delivery_analysis['on_time_delivery_rate'] > 60:
            insights.append(f"⚠️ Средняя производительность доставки: {delivery_analysis['on_time_delivery_rate']:.1f}% заказов доставлено вовремя")
        else:
            insights.append(f"❌ Низкая производительность доставки: {delivery_analysis['on_time_delivery_rate']:.1f}% заказов доставлено вовремя")
        
        # Инсайты по скидкам
        if discount_analysis['discount_percentage'] > 50:
            insights.append(f"🎯 Высокая активность скидок: {discount_analysis['discount_percentage']:.1f}% заказов со скидкой")
        
        # Инсайты по отменам
        if basic_metrics['cancellation_rate'] > 10:
            insights.append(f"⚠️ Высокий процент отмен: {basic_metrics['cancellation_rate']:.1f}%")
        elif basic_metrics['cancellation_rate'] < 5:
            insights.append(f"✅ Низкий процент отмен: {basic_metrics['cancellation_rate']:.1f}%")
        
        # Инсайты по весу
        if weight_analysis.get('weight_data_available', False) and weight_analysis['heavy_orders_percentage'] > 30:
            insights.append(f"📦 Много тяжелых заказов: {weight_analysis['heavy_orders_percentage']:.1f}% заказов >2кг")
        
        return insights